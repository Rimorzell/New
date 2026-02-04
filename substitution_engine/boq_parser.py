"""
BOQ (Bill of Quantities) Parser with Ditto/Stateful Memory Support.
Handles messy, real-world input formats.
"""

import csv
import re
from pathlib import Path
from typing import List, Dict, Optional, Any, Tuple
from dataclasses import dataclass

from .models import (
    BOQItem, IPRating, FormFactor, EnvironmentContext,
    ENVIRONMENT_KEYWORDS, FORM_FACTOR_KEYWORDS, ENVIRONMENT_IP_REQUIREMENTS
)


@dataclass
class ParsedSpecs:
    """Extracted specifications from a text description."""
    wattage: Optional[float] = None
    lumens: Optional[float] = None
    ip_rating: Optional[IPRating] = None
    form_factor: Optional[FormFactor] = None
    environment: EnvironmentContext = EnvironmentContext.INDOOR_DRY
    is_emergency: bool = False
    is_dali: bool = False
    is_dimming: bool = False
    length_mm: Optional[float] = None
    diameter_mm: Optional[float] = None
    cct_k: Optional[float] = None
    beam_deg: Optional[float] = None


class BOQParser:
    """
    Parses customer BOQ files and extracts structured specifications.
    Supports "DITTO" / "Same as above" functionality with stateful memory.
    """

    # Patterns for extracting specifications
    WATTAGE_PATTERNS = [
        r'(\d+(?:\.\d+)?)\s*[wW](?:att)?(?:s)?(?!\s*/)',  # 38W, 38 watt, 38watts
        r'(\d+(?:\.\d+)?)\s*[wW]/[mM]',  # 15W/m for linear
    ]

    LUMEN_PATTERNS = [
        r'(\d+(?:\.\d+)?)\s*(?:lm|lumen|lumens)',  # 3000lm, 3000 lumens
        r'(\d+(?:\.\d+)?)\s*(?:LM|LUMEN|LUMENS)',
    ]

    IP_PATTERN = r'IP\s*(\d)(\d)'  # IP20, IP65, IP 65

    LENGTH_PATTERNS = [
        r'(\d+(?:\.\d+)?)\s*(?:mm|MM)',  # 1200mm
        r'(\d+(?:\.\d+)?)\s*[mM](?!\w)',  # 1.2m (not mm)
        r'L\s*(\d+)',  # L1200
    ]

    CCT_PATTERNS = [
        r'(\d{4})\s*[kK]',  # 4000K
        r'(\d+)\s*kelvin',  # 4000 kelvin
    ]

    DITTO_PATTERNS = [
        r'\bditto\b',
        r'\bsame\s+as\s+above\b',
        r'\bsame\s+as\s+previous\b',
        r'\bas\s+above\b',
        r'\brepeat\b',
        r'\bid\.\b',  # id. (idem)
        r'\bidem\b',
    ]

    BEAM_PATTERNS = [
        r'(\d+(?:\.\d+)?)\s*(?:deg|degree|degrees|°)',
    ]

    def __init__(self):
        self._previous_item: Optional[BOQItem] = None
        self._items: List[BOQItem] = []
        self._row_lookup: Dict[int, BOQItem] = {}

    def parse_file(self, filepath: str) -> List[BOQItem]:
        """
        Parse a BOQ file (CSV or Excel-like).
        Returns list of BOQItem objects.
        """
        filepath = Path(filepath)

        if not filepath.exists():
            raise FileNotFoundError(f"BOQ file not found: {filepath}")

        ext = filepath.suffix.lower()

        if ext == '.csv':
            return self._parse_csv(filepath)
        elif ext in ['.xlsx', '.xls']:
            return self._parse_excel(filepath)
        else:
            # Try CSV parsing as default
            return self._parse_csv(filepath)

    def _parse_csv(self, filepath: Path) -> List[BOQItem]:
        """Parse a CSV file."""
        self._items = []
        self._previous_item = None
        self._row_lookup = {}

        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            # Try to detect delimiter
            sample = f.read(2048)
            f.seek(0)

            # Detect delimiter
            if '\t' in sample:
                delimiter = '\t'
            elif ';' in sample:
                delimiter = ';'
            else:
                delimiter = ','

            reader = csv.DictReader(f, delimiter=delimiter)

            for row_num, row in enumerate(reader, start=2):  # Start at 2 (1 for header)
                item = self._parse_row(row_num, row)
                if item:
                    self._items.append(item)
                    self._row_lookup[row_num] = item
                    self._previous_item = item

        return self._items

    def _parse_excel(self, filepath: Path) -> List[BOQItem]:
        """Parse an Excel file. Requires openpyxl."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required to parse Excel files. "
                            "Install with: pip install openpyxl")

        self._items = []
        self._previous_item = None
        self._row_lookup = {}

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value or '').strip().lower())

        # Process data rows
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(value or '')

            item = self._parse_row(row_num, row_dict)
            if item:
                self._items.append(item)
                self._row_lookup[row_num] = item
                self._previous_item = item

        return self._items

    def parse_description(self, description: str) -> BOQItem:
        """
        Parse a single description string.
        Useful for ad-hoc queries.
        """
        return self._parse_row(1, {'description': description})

    def _parse_row(self, row_num: int, row: Dict) -> Optional[BOQItem]:
        """Parse a single row into a BOQItem."""
        # Find the description field (various possible names)
        description = self._find_description(row)

        if not description or not description.strip():
            return None

        # Check for DITTO
        is_ditto, ditto_mods = self._check_ditto(description, row)

        if is_ditto and self._previous_item:
            # Create item based on previous item with modifications
            item = self._create_ditto_item(row_num, description, ditto_mods)
        else:
            # Parse fresh specifications
            specs = self._extract_specs(description, row)
            item = BOQItem(
                row_number=row_num,
                raw_description=description,
                quantity=self._extract_quantity(row),
                requested_wattage=specs.wattage,
                requested_lumens=specs.lumens,
                requested_ip=specs.ip_rating,
                requested_form_factor=specs.form_factor,
                requested_cct_k=specs.cct_k,
                requested_length_mm=specs.length_mm,
                requested_beam_deg=specs.beam_deg,
                environment=specs.environment,
                requires_emergency=specs.is_emergency,
                requires_dali=specs.is_dali,
                requires_dimming=specs.is_dimming,
                is_ditto=False,
                parsed_fields=dict(row)
            )

        return item

    def _find_description(self, row: Dict) -> str:
        """Find the description field from various possible column names."""
        description_keys = [
            'description', 'desc', 'item', 'item description',
            'product', 'product description', 'specification',
            'spec', 'details', 'name', 'material', 'material description'
        ]

        # First, try exact matches
        for key in description_keys:
            if key in row and row[key]:
                return str(row[key]).strip()

        # Try case-insensitive matches
        row_lower = {k.lower(): v for k, v in row.items()}
        for key in description_keys:
            if key in row_lower and row_lower[key]:
                return str(row_lower[key]).strip()

        # Concatenate all non-empty fields as fallback
        parts = [str(v).strip() for v in row.values() if v and str(v).strip()]
        return ' | '.join(parts)

    def _check_ditto(self, description: str, row: Dict) -> Tuple[bool, Dict]:
        """
        Check if this row is a DITTO/repeat reference.
        Returns (is_ditto, modifications_dict).
        """
        desc_lower = description.lower()
        modifications = {}

        # Check for ditto patterns
        for pattern in self.DITTO_PATTERNS:
            if re.search(pattern, desc_lower, re.IGNORECASE):
                # Extract any modifications mentioned
                modifications = self._extract_ditto_modifications(description)
                return True, modifications

        # Check for "Same as [reference]" pattern
        same_as_match = re.search(r'same\s+as\s+(\w+\d*)', desc_lower)
        if same_as_match:
            modifications = self._extract_ditto_modifications(description)
            return True, modifications

        return False, {}

    def _extract_ditto_modifications(self, description: str) -> Dict:
        """Extract modifications from a ditto reference (e.g., 'but with emergency')."""
        mods = {}
        desc_lower = description.lower()

        # Check for "but" modifications
        but_match = re.search(r'but\s+(.+)', desc_lower)
        if but_match:
            mod_text = but_match.group(1)

            # Check for emergency
            if 'emergency' in mod_text or 'em' in mod_text.split():
                mods['requires_emergency'] = True

            # Check for DALI
            if 'dali' in mod_text:
                mods['requires_dali'] = True

            # Check for IP change
            ip_match = re.search(r'IP\s*(\d)(\d)', mod_text, re.IGNORECASE)
            if ip_match:
                mods['requested_ip'] = IPRating(
                    int(ip_match.group(1)),
                    int(ip_match.group(2)),
                    f"IP{ip_match.group(1)}{ip_match.group(2)}"
                )

            # Check for wattage change
            wattage_match = re.search(r'(\d+)\s*[wW]', mod_text)
            if wattage_match:
                mods['requested_wattage'] = float(wattage_match.group(1))

        return mods

    def _create_ditto_item(self, row_num: int, description: str,
                          modifications: Dict) -> BOQItem:
        """Create a new item based on the previous item with modifications."""
        prev = self._previous_item

        item = BOQItem(
            row_number=row_num,
            raw_description=description,
            quantity=1,
            requested_wattage=modifications.get('requested_wattage', prev.requested_wattage),
            requested_lumens=modifications.get('requested_lumens', prev.requested_lumens),
            requested_ip=modifications.get('requested_ip', prev.requested_ip),
            requested_form_factor=modifications.get('requested_form_factor', prev.requested_form_factor),
            requested_cct_k=modifications.get('requested_cct_k', prev.requested_cct_k),
            requested_length_mm=modifications.get('requested_length_mm', prev.requested_length_mm),
            requested_beam_deg=modifications.get('requested_beam_deg', prev.requested_beam_deg),
            environment=modifications.get('environment', prev.environment),
            requires_emergency=modifications.get('requires_emergency', prev.requires_emergency),
            requires_dali=modifications.get('requires_dali', prev.requires_dali),
            requires_dimming=modifications.get('requires_dimming', prev.requires_dimming),
            is_ditto=True,
            ditto_source_row=prev.row_number,
            ditto_modifications=modifications
        )

        return item

    def _extract_specs(self, description: str, row: Dict) -> ParsedSpecs:
        """Extract all specifications from description and row fields."""
        specs = ParsedSpecs()
        combined_text = f"{description} {' '.join(str(v) for v in row.values())}"

        # Extract wattage
        specs.wattage = self._extract_wattage(combined_text)

        # Extract lumens
        specs.lumens = self._extract_lumens(combined_text)

        # Extract IP rating
        specs.ip_rating = self._extract_ip_rating(combined_text)

        # Infer environment context
        specs.environment = self._infer_environment(combined_text)

        # If IP not specified but environment suggests higher IP, set it
        if specs.ip_rating is None and specs.environment != EnvironmentContext.INDOOR_DRY:
            specs.ip_rating = ENVIRONMENT_IP_REQUIREMENTS.get(specs.environment)

        # Extract form factor
        specs.form_factor = self._infer_form_factor(combined_text)

        # Extract features
        specs.is_emergency = self._check_emergency(combined_text)
        specs.is_dali = self._check_dali(combined_text)
        specs.is_dimming = self._check_dimming(combined_text)

        # Extract CCT
        specs.cct_k = self._extract_cct(combined_text)

        # Extract dimensions
        specs.length_mm = self._extract_length(combined_text)
        specs.beam_deg = self._extract_beam(combined_text)

        return specs

    def _extract_wattage(self, text: str) -> Optional[float]:
        """Extract wattage from text."""
        for pattern in self.WATTAGE_PATTERNS:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return float(match.group(1))
        return None

    def _extract_lumens(self, text: str) -> Optional[float]:
        """Extract lumens from text."""
        for pattern in self.LUMEN_PATTERNS:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return float(match.group(1))
        return None

    def _extract_ip_rating(self, text: str) -> Optional[IPRating]:
        """Extract IP rating from text."""
        match = re.search(self.IP_PATTERN, text, re.IGNORECASE)
        if match:
            return IPRating(
                int(match.group(1)),
                int(match.group(2)),
                f"IP{match.group(1)}{match.group(2)}"
            )
        return None

    def _extract_cct(self, text: str) -> Optional[float]:
        """Extract color temperature from text."""
        for pattern in self.CCT_PATTERNS:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return float(match.group(1))
        return None

    def _extract_length(self, text: str) -> Optional[float]:
        """Extract length from text."""
        for pattern in self.LENGTH_PATTERNS:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                value = float(match.group(1))
                # Convert meters to mm if needed
                if 'm' in pattern.lower() and value < 10:
                    value *= 1000
                return value
        return None

    def _extract_beam(self, text: str) -> Optional[float]:
        """Extract beam angle in degrees from text."""
        for pattern in self.BEAM_PATTERNS:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return float(match.group(1))
        return None

    def _infer_environment(self, text: str) -> EnvironmentContext:
        """Infer environment context from text."""
        text_lower = text.lower()

        # Check each environment type
        for env, keywords in ENVIRONMENT_KEYWORDS.items():
            for keyword in keywords:
                if keyword in text_lower:
                    return env

        return EnvironmentContext.INDOOR_DRY

    def _infer_form_factor(self, text: str) -> Optional[FormFactor]:
        """Infer form factor from text."""
        text_lower = text.lower()

        # Score each form factor by keyword matches
        scores = {}
        for form_factor, keywords in FORM_FACTOR_KEYWORDS.items():
            score = 0
            for keyword in keywords:
                if keyword in text_lower:
                    # Weight longer keywords higher
                    score += len(keyword.split())
            if score > 0:
                scores[form_factor] = score

        if scores:
            return max(scores.keys(), key=lambda k: scores[k])

        return None

    def _check_emergency(self, text: str) -> bool:
        """Check if emergency lighting is required."""
        patterns = [
            r'\bemergency\b',
            r'\bEM\b',
            r'\bself[-\s]?contained\b',
            r'\bbattery\s+backup\b',
            r'\b3\s*hr\b',
            r'\b3\s*hour\b',
        ]
        for pattern in patterns:
            if re.search(pattern, text, re.IGNORECASE):
                return True
        return False

    def _check_dali(self, text: str) -> bool:
        """Check if DALI is required."""
        return bool(re.search(r'\bDALI\b', text, re.IGNORECASE))

    def _check_dimming(self, text: str) -> bool:
        """Check if dimming is required."""
        patterns = [
            r'\bdimmable\b',
            r'\bdimming\b',
            r'\bdim\b',
            r'\b1-10v\b',
            r'\b0-10v\b',
            r'\btriac\b',
            r'\bphase[-\s]?cut\b',
        ]
        for pattern in patterns:
            if re.search(pattern, text, re.IGNORECASE):
                return True
        return False

    def _extract_quantity(self, row: Dict) -> int:
        """Extract quantity from row."""
        qty_keys = ['qty', 'quantity', 'count', 'amount', 'no', 'units']

        for key in qty_keys:
            if key in row:
                try:
                    return int(float(row[key]))
                except (ValueError, TypeError):
                    pass

            # Try case-insensitive
            for row_key, value in row.items():
                if row_key.lower() == key:
                    try:
                        return int(float(value))
                    except (ValueError, TypeError):
                        pass

        return 1
